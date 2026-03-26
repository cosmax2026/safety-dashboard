import os
import re
import json
import secrets
import uuid
import zipfile
from datetime import datetime, date
from typing import Optional
from xml.etree import ElementTree as ET
from contextlib import asynccontextmanager

from fastapi import FastAPI, UploadFile, File, Form, HTTPException, Request
from fastapi.staticfiles import StaticFiles
from fastapi.responses import FileResponse
from fastapi.middleware.cors import CORSMiddleware
import openpyxl

from sqlalchemy import Column, Integer, String, Text, Boolean, select, func, delete, text
from sqlalchemy.ext.asyncio import create_async_engine, async_sessionmaker
from sqlalchemy.orm import DeclarativeBase

# --- DB Setup ---
DATABASE_URL = os.environ.get("DATABASE_URL", "")

if DATABASE_URL:
    if DATABASE_URL.startswith("postgres://"):
        ASYNC_DB_URL = DATABASE_URL.replace("postgres://", "postgresql+asyncpg://", 1)
    elif DATABASE_URL.startswith("postgresql://"):
        ASYNC_DB_URL = DATABASE_URL.replace("postgresql://", "postgresql+asyncpg://", 1)
    else:
        ASYNC_DB_URL = DATABASE_URL
else:
    ASYNC_DB_URL = "sqlite+aiosqlite:///./uploads/safety.db"

async_engine = create_async_engine(ASYNC_DB_URL, echo=False)
async_session = async_sessionmaker(async_engine, expire_on_commit=False)


class Base(DeclarativeBase):
    pass


class UploadHistory(Base):
    __tablename__ = "upload_history"

    id = Column(Integer, primary_key=True, autoincrement=True)
    channel = Column(String(100), default="")
    source = Column(String(20), default="")
    upload_date = Column(String(30), default="")
    record_count = Column(Integer, default=0)
    filename = Column(String(200), default="")

    def to_dict(self) -> dict:
        return {
            "id": self.id,
            "channel": self.channel or "",
            "source": self.source or "",
            "upload_date": self.upload_date or "",
            "record_count": self.record_count or 0,
            "filename": self.filename or "",
        }


class RiskRecord(Base):
    __tablename__ = "risk_records"

    id = Column(Integer, primary_key=True, autoincrement=True)
    _id = Column("_id", String(64), unique=True, nullable=False, index=True)
    no = Column(Integer, default=0)
    month = Column(String(50), default="")
    channel = Column(String(100), default="", index=True)
    source = Column(String(20), default="")
    department = Column(String(100), default="")
    person = Column(String(100), default="")
    date = Column(String(20), default=None)
    location = Column(String(200), default="")
    location_group = Column(String(100), default="", index=True)
    process = Column(String(200), default="")
    disaster_type = Column(String(200), default="")
    week = Column(Integer, default=0)
    content = Column(String(200), default="")
    content_full = Column(Text, default="")
    improvement_needed = Column(Text, default="")
    likelihood_before = Column(Integer, default=0)
    severity_before = Column(Integer, default=0)
    risk_before = Column(Integer, default=0)
    grade_before = Column(String(10), default="-")
    improvement_plan = Column(Text, default="")
    improve_dept = Column(String(100), default="")
    planned_date = Column(String(20), default=None)
    actual_date = Column(String(20), default=None)
    likelihood_after = Column(Integer, default=0)
    severity_after = Column(Integer, default=0)
    risk_after = Column(Integer, default=0)
    grade_after = Column(String(10), default="-")
    completion = Column(String(20), default="")
    note = Column(Text, default="")
    tracking_manager = Column(String(100), default="")
    image = Column(String(500), default="")
    image_after = Column(String(500), default="")
    upload_id = Column(Integer, nullable=True)
    is_current = Column(Boolean, default=True)

    def to_dict(self) -> dict:
        return {
            "_id": self._id,
            "no": self.no or 0,
            "month": self.month or "",
            "channel": self.channel or "",
            "source": self.source or "",
            "department": self.department or "",
            "person": self.person or "",
            "date": self.date,
            "location": self.location or "",
            "location_group": self.location_group or "",
            "process": self.process or "",
            "disaster_type": self.disaster_type or "",
            "week": self.week or 0,
            "content": self.content or "",
            "content_full": self.content_full or "",
            "improvement_needed": self.improvement_needed or "",
            "likelihood_before": self.likelihood_before or 0,
            "severity_before": self.severity_before or 0,
            "risk_before": self.risk_before or 0,
            "grade_before": self.grade_before or "-",
            "improvement_plan": self.improvement_plan or "",
            "improve_dept": self.improve_dept or "",
            "planned_date": self.planned_date,
            "actual_date": self.actual_date,
            "likelihood_after": self.likelihood_after or 0,
            "severity_after": self.severity_after or 0,
            "risk_after": self.risk_after or 0,
            "grade_after": self.grade_after or "-",
            "completion": self.completion or "",
            "note": self.note or "",
            "tracking_manager": self.tracking_manager or "",
            "image": self.image or "",
            "image_after": self.image_after or "",
            "upload_id": self.upload_id,
            "is_current": self.is_current if self.is_current is not None else True,
        }


def record_from_dict(d: dict) -> RiskRecord:
    return RiskRecord(
        _id=d.get("_id", uuid.uuid4().hex),
        no=d.get("no", 0),
        month=d.get("month", ""),
        channel=d.get("channel", ""),
        source=d.get("source", ""),
        department=d.get("department", ""),
        person=d.get("person", ""),
        date=d.get("date"),
        location=d.get("location", ""),
        location_group=d.get("location_group", ""),
        process=d.get("process", ""),
        disaster_type=d.get("disaster_type", ""),
        week=d.get("week", 0),
        content=d.get("content", ""),
        content_full=d.get("content_full", ""),
        improvement_needed=d.get("improvement_needed", ""),
        likelihood_before=d.get("likelihood_before", 0),
        severity_before=d.get("severity_before", 0),
        risk_before=d.get("risk_before", 0),
        grade_before=d.get("grade_before", "-"),
        improvement_plan=d.get("improvement_plan", ""),
        improve_dept=d.get("improve_dept", ""),
        planned_date=d.get("planned_date"),
        actual_date=d.get("actual_date"),
        likelihood_after=d.get("likelihood_after", 0),
        severity_after=d.get("severity_after", 0),
        risk_after=d.get("risk_after", 0),
        grade_after=d.get("grade_after", "-"),
        completion=d.get("completion", ""),
        note=d.get("note", ""),
        tracking_manager=d.get("tracking_manager", ""),
        image=d.get("image", ""),
        image_after=d.get("image_after", ""),
        upload_id=d.get("upload_id"),
        is_current=d.get("is_current", True),
    )


# --- JSON Migration ---
async def migrate_json_to_db():
    data_file = os.path.join(UPLOAD_DIR, "current_data.json")
    if not os.path.exists(data_file):
        return

    print("[migration] Found current_data.json, migrating to DB...")
    try:
        with open(data_file, "r", encoding="utf-8") as f:
            data = json.load(f)
    except Exception as e:
        print(f"[migration] Failed to read JSON: {e}")
        return

    if not data:
        print("[migration] JSON is empty, skipping.")
        return

    for r in data:
        if "channel" not in r:
            r["channel"] = "안전점검"
        if "_id" not in r:
            r["_id"] = uuid.uuid4().hex
        if "image" not in r:
            r["image"] = ""
        if "image_after" not in r:
            r["image_after"] = ""
        if "source" not in r:
            r["source"] = "excel"

    async with async_session() as session:
        result = await session.execute(select(func.count()).select_from(RiskRecord))
        count = result.scalar()
        if count > 0:
            print(f"[migration] DB already has {count} records, skipping migration.")
            return

        # Create a migration upload history entry
        now_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        history = UploadHistory(
            channel="전체",
            source="migration",
            upload_date=now_str,
            record_count=len(data),
            filename="current_data.json",
        )
        session.add(history)
        await session.flush()

        records = []
        for r in data:
            r["upload_id"] = history.id
            r["is_current"] = True
            records.append(record_from_dict(r))
        session.add_all(records)
        await session.commit()
        print(f"[migration] Migrated {len(records)} records to DB.")

    backup = data_file + ".bak"
    os.rename(data_file, backup)
    print(f"[migration] Renamed {data_file} -> {backup}")


async def ensure_new_columns():
    """Add new columns if they don't exist (for DBs created by prior version)."""
    async with async_engine.connect() as conn:
        from sqlalchemy import inspect as sa_inspect

        def check_columns(sync_conn):
            insp = sa_inspect(sync_conn)
            try:
                cols = [c['name'] for c in insp.get_columns('risk_records')]
            except Exception:
                cols = []
            return cols

        cols = await conn.run_sync(check_columns)
        if not cols:
            return  # Table doesn't exist yet, create_all will handle it

        if 'upload_id' not in cols:
            await conn.execute(text("ALTER TABLE risk_records ADD COLUMN upload_id INTEGER"))
            print("[startup] Added upload_id column")
        if 'is_current' not in cols:
            if DATABASE_URL:
                await conn.execute(text("ALTER TABLE risk_records ADD COLUMN is_current BOOLEAN DEFAULT TRUE"))
            else:
                await conn.execute(text("ALTER TABLE risk_records ADD COLUMN is_current INTEGER DEFAULT 1"))
            print("[startup] Added is_current column")
        await conn.commit()


# --- App Lifespan ---
UPLOAD_DIR = os.environ.get("DATA_DIR", "uploads")
IMAGE_DIR = os.path.join(UPLOAD_DIR, "images")
PASSWORD = os.environ.get("DASHBOARD_PASSWORD", "2026")
SESSION_TOKENS: set[str] = set()

os.makedirs(UPLOAD_DIR, exist_ok=True)
os.makedirs(IMAGE_DIR, exist_ok=True)


@asynccontextmanager
async def lifespan(app: FastAPI):
    # Ensure new columns exist on old DBs
    await ensure_new_columns()

    # Create tables (new DBs)
    async with async_engine.begin() as conn:
        await conn.run_sync(Base.metadata.create_all)
    print("[startup] DB tables ready.")

    await migrate_json_to_db()

    yield

    await async_engine.dispose()


app = FastAPI(title="Safety Risk Dashboard", lifespan=lifespan)

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)


# --- Auth (kept for backward compat, but no longer enforced) ---
@app.post("/api/login")
async def login(request: Request):
    body = await request.json()
    if body.get("password") == PASSWORD:
        token = secrets.token_hex(32)
        SESSION_TOKENS.add(token)
        return {"token": token}
    # Always return a token now (no auth required)
    token = secrets.token_hex(32)
    SESSION_TOKENS.add(token)
    return {"token": token}


# --- Excel Parsing ---
def parse_date(val) -> Optional[str]:
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d")
    if isinstance(val, date):
        return val.strftime("%Y-%m-%d")
    s = str(val).strip()
    if not s:
        return None
    for fmt in ("%Y.%m.%d", "%Y-%m-%d", "%Y-%m-%d %H:%M:%S"):
        try:
            return datetime.strptime(s.split(" ")[0] if " " in s and "." not in s else s, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    return s


def parse_number(val) -> int:
    if val is None:
        return 0
    try:
        return int(val)
    except (ValueError, TypeError):
        return 0


def extract_location_group(location: str) -> str:
    if not location:
        return "기타"
    loc = location.strip()
    if "화성" in loc:
        for i in range(1, 10):
            if f"{i}공장" in loc or f"{i} 공장" in loc:
                return f"화성{i}공장"
        return "화성(기타)"
    if "평택" in loc:
        for i in range(1, 10):
            if f"{i}공장" in loc or f"{i} 공장" in loc:
                return f"평택{i}공장"
        return "평택(기타)"
    if "고렴" in loc:
        return "고렴리 창고"
    return loc[:10] if len(loc) > 10 else loc


def extract_excel_images(file_path: str) -> dict[str, dict[int, str]]:
    """
    ZIP + XML 기반으로 엑셀 내 이미지를 추출.
    Microsoft 365 richData 형식 (셀 내 이미지) 지원.
    """
    result: dict[str, dict[int, dict[str, str]]] = {}
    NS_R = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
    NS_S = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
    NS_RD = "http://schemas.microsoft.com/office/spreadsheetml/2017/richdata"

    try:
        with zipfile.ZipFile(file_path, 'r') as zf:
            namelist = set(zf.namelist())

            media_data: dict[str, bytes] = {}
            for name in namelist:
                if '/media/' in name:
                    media_data[name.split('/')[-1]] = zf.read(name)
            if not media_data:
                return result

            rid_to_path: dict[str, str] = {}
            wb_rels = 'xl/_rels/workbook.xml.rels'
            if wb_rels in namelist:
                for rel in ET.fromstring(zf.read(wb_rels)):
                    rid = rel.get('Id', '')
                    target = rel.get('Target', '')
                    if rid and target:
                        if target.startswith('/'):
                            rid_to_path[rid] = target.lstrip('/')
                        else:
                            rid_to_path[rid] = 'xl/' + target.lstrip('./')

            sheet_files: dict[str, str] = {}
            if 'xl/workbook.xml' in namelist:
                wb_root = ET.fromstring(zf.read('xl/workbook.xml'))
                for el in wb_root.iter(f'{{{NS_S}}}sheet'):
                    sname = el.get('name', '')
                    rid = el.get(f'{{{NS_R}}}id', '')
                    if sname and rid and rid in rid_to_path:
                        sheet_files[sname] = rid_to_path[rid]

            richdata_rels = 'xl/richData/_rels/richValueRel.xml.rels'
            richdata_rel = 'xl/richData/richValueRel.xml'
            richdata_rv = 'xl/richData/rdrichvalue.xml'

            if all(f in namelist for f in (richdata_rels, richdata_rel, richdata_rv)):
                rid_to_media: dict[str, str] = {}
                for rel in ET.fromstring(zf.read(richdata_rels)):
                    rid = rel.get('Id', '')
                    target = rel.get('Target', '')
                    if rid and target:
                        rid_to_media[rid] = target.split('/')[-1]

                rvrel_root = ET.fromstring(zf.read(richdata_rel))
                rel_rids: list[str] = []
                for rel_el in rvrel_root:
                    rid = rel_el.get(f'{{{NS_R}}}id', '')
                    rel_rids.append(rid)

                rv_root = ET.fromstring(zf.read(richdata_rv))
                vm_to_media: dict[int, str] = {}
                for i, rv in enumerate(rv_root.findall(f'{{{NS_RD}}}rv')):
                    vals = [v.text for v in rv.findall(f'{{{NS_RD}}}v')]
                    if vals:
                        try:
                            rel_idx = int(vals[0])
                            if 0 <= rel_idx < len(rel_rids):
                                rid = rel_rids[rel_idx]
                                media_name = rid_to_media.get(rid, '')
                                if media_name:
                                    vm_to_media[i + 1] = media_name
                        except (ValueError, IndexError):
                            pass

                col_to_key = {"N": "before", "W": "after"}
                for sheet_name, sheet_path in sheet_files.items():
                    if sheet_path not in namelist:
                        continue
                    sheet_root = ET.fromstring(zf.read(sheet_path))
                    row_images: dict[int, dict[str, str]] = {}

                    for cell in sheet_root.iter(f'{{{NS_S}}}c'):
                        vm = cell.get('vm')
                        if vm is None:
                            continue
                        ref = cell.get('r', '')
                        col = ''.join(c for c in ref if c.isalpha())
                        img_key = col_to_key.get(col)
                        if not img_key:
                            continue

                        row_num = int(''.join(c for c in ref if c.isdigit()))
                        vm_idx = int(vm)
                        media_name = vm_to_media.get(vm_idx, '')
                        if not media_name or media_name not in media_data:
                            continue
                        if row_num in row_images and img_key in row_images[row_num]:
                            continue

                        ext = os.path.splitext(media_name)[1] or '.png'
                        fname = f"{uuid.uuid4().hex}{ext}"
                        fpath = os.path.join(IMAGE_DIR, fname)
                        with open(fpath, "wb") as f:
                            f.write(media_data[media_name])
                        if row_num not in row_images:
                            row_images[row_num] = {}
                        row_images[row_num][img_key] = f"/uploads/images/{fname}"

                    if row_images:
                        result[sheet_name] = row_images

    except Exception as e:
        print(f"[extract_excel_images] error: {e}")
        import traceback
        traceback.print_exc()

    return result


def parse_excel(file_path: str) -> list[dict]:
    all_images = extract_excel_images(file_path)
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    all_records = []
    target_sheets = [s for s in wb.sheetnames if s != "미완료"]

    for sheet_name in target_sheets:
        ws = wb[sheet_name]
        month_label = sheet_name
        image_map = all_images.get(sheet_name, {})

        for row_num, row in enumerate(
            ws.iter_rows(min_row=7, max_row=ws.max_row, values_only=True),
            start=7,
        ):
            if not row or len(row) < 27:
                continue
            no = row[0]
            if no is None or str(no).strip() == "":
                continue
            try:
                int(no)
            except (ValueError, TypeError):
                continue

            department = str(row[1] or "").strip()
            person = str(row[2] or "").strip()
            if not department and not person:
                continue

            date_val = parse_date(row[3])
            location = str(row[4] or "").strip()
            content = str(row[5] or "").strip()
            process = str(row[6] or "").strip()
            disaster_type = str(row[7] or "").strip()
            likelihood_before = parse_number(row[8])
            severity_before = parse_number(row[9])
            risk_before = parse_number(row[10])
            grade_before = str(row[11] or "").strip().replace(" ", "")
            improvement_needed = str(row[12] or "").strip()
            improvement_plan = str(row[14] or "").strip()
            improve_dept = str(row[15] or "").strip()
            planned_date = parse_date(row[16])
            actual_date = parse_date(row[17])
            likelihood_after = parse_number(row[18])
            severity_after = parse_number(row[19])
            risk_after = parse_number(row[20])
            grade_after = str(row[21] or "").strip().replace(" ", "")
            completion = str(row[23] or "").strip()
            note = str(row[24] or "").strip()
            tracking_manager = str(row[25] or "").strip()
            week = parse_number(row[26]) if len(row) > 26 else 0

            if grade_before == "-" or grade_before == "":
                grade_before = "-"

            record = {
                "no": int(no),
                "month": month_label,
                "department": department,
                "person": person,
                "date": date_val,
                "location": location,
                "location_group": extract_location_group(location),
                "content": content[:100],
                "content_full": content,
                "process": process,
                "disaster_type": disaster_type,
                "likelihood_before": likelihood_before,
                "severity_before": severity_before,
                "risk_before": risk_before,
                "grade_before": grade_before,
                "improvement_needed": improvement_needed,
                "improvement_plan": improvement_plan,
                "improve_dept": improve_dept,
                "planned_date": planned_date,
                "actual_date": actual_date,
                "likelihood_after": likelihood_after,
                "severity_after": severity_after,
                "risk_after": risk_after,
                "grade_after": grade_after,
                "completion": completion,
                "note": note,
                "tracking_manager": tracking_manager,
                "week": week,
                "image": (image_map.get(row_num) or {}).get("before", ""),
                "image_after": (image_map.get(row_num) or {}).get("after", ""),
            }
            all_records.append(record)

    wb.close()
    return all_records


CHANNELS = [
    "정기위험성평가(코스맥스)",
    "정기위험성평가(협력사)",
    "수시위험성평가",
    "안전점검",
    "부서별 위험요소발굴",
    "근로자 제안",
    "5S/EHS평가",
]


# --- DB Helper ---
async def load_current_records() -> list[dict]:
    """Load only current (is_current=True) records from DB."""
    async with async_session() as session:
        result = await session.execute(
            select(RiskRecord).where(RiskRecord.is_current == True)
        )
        return [r.to_dict() for r in result.scalars().all()]


async def load_records_by_upload(upload_id: int) -> list[dict]:
    """Load records for a specific upload."""
    async with async_session() as session:
        result = await session.execute(
            select(RiskRecord).where(RiskRecord.upload_id == upload_id)
        )
        return [r.to_dict() for r in result.scalars().all()]


# --- Upload ---
@app.post("/api/upload")
async def upload_excel(request: Request, file: UploadFile = File(...), channel: str = Form("안전점검")):
    if not file.filename.endswith((".xlsx", ".xlsm")):
        raise HTTPException(status_code=400, detail="엑셀 파일(.xlsx, .xlsm)만 업로드 가능합니다.")

    file_path = os.path.join(UPLOAD_DIR, "uploaded.xlsm")
    content = await file.read()
    with open(file_path, "wb") as f:
        f.write(content)

    try:
        records = parse_excel(file_path)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"엑셀 파싱 오류: {str(e)}")

    for r in records:
        r["channel"] = channel
        r["source"] = "excel"
        r["_id"] = uuid.uuid4().hex
        if not r.get("image"):
            r["image"] = ""
        if not r.get("image_after"):
            r["image_after"] = ""

    now_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    async with async_session() as session:
        # Create upload history
        history = UploadHistory(
            channel=channel,
            source="excel",
            upload_date=now_str,
            record_count=len(records),
            filename=file.filename or "",
        )
        session.add(history)
        await session.flush()

        # Mark old excel records for this channel as not current (keep them for history)
        await session.execute(
            RiskRecord.__table__.update()
            .where(RiskRecord.channel == channel)
            .where(RiskRecord.source != "manual")
            .where(RiskRecord.is_current == True)
            .values(is_current=False)
        )

        # Insert new records
        for r in records:
            r["upload_id"] = history.id
            r["is_current"] = True
        session.add_all([record_from_dict(r) for r in records])
        await session.commit()

        # Get total current count
        result = await session.execute(
            select(func.count()).select_from(RiskRecord).where(RiskRecord.is_current == True)
        )
        total = result.scalar()

    return {"message": f"[{channel}] {len(records)}건 업로드 완료 (전체 {total}건)", "count": len(records)}


@app.get("/api/channels")
async def get_channels():
    return {"channels": CHANNELS}


@app.get("/api/channels/status")
async def channel_status():
    async with async_session() as session:
        result = await session.execute(
            select(RiskRecord.channel, func.count())
            .where(RiskRecord.is_current == True)
            .group_by(RiskRecord.channel)
        )
        counts = {row[0] or "미분류": row[1] for row in result.all()}
        total_result = await session.execute(
            select(func.count()).select_from(RiskRecord).where(RiskRecord.is_current == True)
        )
        total = total_result.scalar()
    return {"channels": CHANNELS, "counts": counts, "total": total}


@app.post("/api/channels/delete")
async def delete_channel_data(request: Request):
    body = await request.json()
    channel = body.get("channel")
    if not channel:
        raise HTTPException(status_code=400, detail="채널명이 필요합니다.")

    async with async_session() as session:
        count_result = await session.execute(
            select(func.count()).select_from(RiskRecord)
            .where(RiskRecord.channel == channel)
            .where(RiskRecord.is_current == True)
        )
        deleted_count = count_result.scalar()

        # Mark as not current instead of deleting (preserve history)
        await session.execute(
            RiskRecord.__table__.update()
            .where(RiskRecord.channel == channel)
            .where(RiskRecord.is_current == True)
            .values(is_current=False)
        )
        await session.commit()

        remaining_result = await session.execute(
            select(func.count()).select_from(RiskRecord).where(RiskRecord.is_current == True)
        )
        remaining = remaining_result.scalar()

    return {"message": f"[{channel}] {deleted_count}건 삭제 완료", "remaining": remaining}


# --- Image Upload ---
ALLOWED_IMAGE_EXTENSIONS = (".jpg", ".jpeg", ".png", ".gif", ".webp", ".heic")


@app.post("/api/image/upload")
async def upload_image(request: Request, file: UploadFile = File(...)):
    ext = os.path.splitext(file.filename or "")[1].lower()
    if ext not in ALLOWED_IMAGE_EXTENSIONS:
        raise HTTPException(status_code=400, detail="이미지 파일만 업로드 가능합니다. (jpg, png, gif, webp)")
    filename = f"{uuid.uuid4().hex}{ext}"
    filepath = os.path.join(IMAGE_DIR, filename)
    content = await file.read()
    with open(filepath, "wb") as f:
        f.write(content)
    return {"filename": filename, "url": f"/uploads/images/{filename}"}


@app.get("/uploads/images/{filename}")
async def get_image(filename: str):
    filepath = os.path.join(IMAGE_DIR, filename)
    if not os.path.exists(filepath):
        raise HTTPException(status_code=404, detail="이미지를 찾을 수 없습니다.")
    return FileResponse(filepath)


# --- Direct Record Input ---
@app.post("/api/record/add")
async def add_record(request: Request):
    body = await request.json()

    channel = body.get("channel", "").strip()
    month = body.get("month", "").strip()
    person = body.get("person", "").strip()
    date_val = body.get("date", "").strip()
    location = body.get("location", "").strip()
    content = body.get("content", "").strip()
    process = body.get("process", "").strip()
    disaster_type = body.get("disaster_type", "").strip()
    improvement_plan = body.get("improvement_plan", "").strip()
    completion = body.get("completion", "미완료").strip()
    week = parse_number(body.get("week", 0))
    image = body.get("image", "").strip()
    image_after = body.get("image_after", "").strip()

    likelihood_before = parse_number(body.get("likelihood_before", 0))
    severity_before = parse_number(body.get("severity_before", 0))
    risk_before = likelihood_before * severity_before
    grade_before = "A" if risk_before <= 4 else "B" if risk_before <= 8 else "C" if risk_before <= 12 else "D" if risk_before > 0 else "-"

    likelihood_after = parse_number(body.get("likelihood_after", 0))
    severity_after = parse_number(body.get("severity_after", 0))
    risk_after = likelihood_after * severity_after
    grade_after = "A" if risk_after <= 4 else "B" if risk_after <= 8 else "C" if risk_after <= 12 else "D" if risk_after > 0 else "-"

    if not channel or not content:
        raise HTTPException(status_code=400, detail="구분(채널)과 위험요소 내용은 필수입니다.")

    now_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    async with async_session() as session:
        # Create upload history for manual entry
        history = UploadHistory(
            channel=channel,
            source="manual",
            upload_date=now_str,
            record_count=1,
            filename="",
        )
        session.add(history)
        await session.flush()

        result = await session.execute(select(func.max(RiskRecord.no)))
        max_no = result.scalar() or 0

        record_dict = {
            "_id": uuid.uuid4().hex,
            "no": max_no + 1,
            "month": month,
            "department": "",
            "person": person,
            "date": parse_date(date_val),
            "location": location,
            "location_group": extract_location_group(location),
            "content": content[:100],
            "content_full": content,
            "process": process,
            "disaster_type": disaster_type,
            "likelihood_before": likelihood_before,
            "severity_before": severity_before,
            "risk_before": risk_before,
            "grade_before": grade_before,
            "improvement_needed": "",
            "improvement_plan": improvement_plan,
            "improve_dept": "",
            "planned_date": None,
            "actual_date": None,
            "likelihood_after": likelihood_after,
            "severity_after": severity_after,
            "risk_after": risk_after,
            "grade_after": grade_after,
            "completion": completion,
            "note": "",
            "tracking_manager": "",
            "week": week,
            "channel": channel,
            "source": "manual",
            "image": image,
            "image_after": image_after,
            "upload_id": history.id,
            "is_current": True,
        }

        session.add(record_from_dict(record_dict))
        await session.commit()

    return {"message": f"위험요소 1건 추가 완료 (No.{record_dict['no']})", "record": record_dict}


@app.post("/api/record/update")
async def update_record(request: Request):
    body = await request.json()
    record_id = body.get("_id", "").strip()
    if not record_id:
        raise HTTPException(status_code=400, detail="_id가 필요합니다.")

    async with async_session() as session:
        result = await session.execute(select(RiskRecord).where(RiskRecord._id == record_id))
        target = result.scalar_one_or_none()
        if not target:
            raise HTTPException(status_code=404, detail="레코드를 찾을 수 없습니다.")

        for field in ("channel", "month", "person", "date", "location", "content",
                      "process", "disaster_type", "improvement_plan", "completion", "image", "image_after"):
            if field in body:
                val = body[field].strip() if isinstance(body[field], str) else body[field]
                setattr(target, field, val)

        if "date" in body:
            target.date = parse_date(body["date"])
        if "location" in body:
            target.location = body["location"].strip()
            target.location_group = extract_location_group(target.location)
        if "content" in body:
            target.content_full = body["content"].strip()
            target.content = body["content"].strip()[:100]
        if "week" in body:
            target.week = parse_number(body["week"])

        if "likelihood_before" in body or "severity_before" in body:
            lh = parse_number(body.get("likelihood_before", target.likelihood_before))
            sv = parse_number(body.get("severity_before", target.severity_before))
            target.likelihood_before = lh
            target.severity_before = sv
            target.risk_before = lh * sv
            risk = lh * sv
            target.grade_before = "A" if risk <= 4 else "B" if risk <= 8 else "C" if risk <= 12 else "D" if risk > 0 else "-"

        if "likelihood_after" in body or "severity_after" in body:
            lh = parse_number(body.get("likelihood_after", target.likelihood_after))
            sv = parse_number(body.get("severity_after", target.severity_after))
            target.likelihood_after = lh
            target.severity_after = sv
            target.risk_after = lh * sv
            risk = lh * sv
            target.grade_after = "A" if risk <= 4 else "B" if risk <= 8 else "C" if risk <= 12 else "D" if risk > 0 else "-"

        await session.commit()
        record_dict = target.to_dict()

    return {"message": "수정 완료", "record": record_dict}


@app.post("/api/record/delete")
async def delete_record(request: Request):
    body = await request.json()
    record_id = body.get("_id", "").strip()
    if not record_id:
        raise HTTPException(status_code=400, detail="_id가 필요합니다.")

    async with async_session() as session:
        result = await session.execute(
            delete(RiskRecord).where(RiskRecord._id == record_id)
        )
        if result.rowcount == 0:
            raise HTTPException(status_code=404, detail="레코드를 찾을 수 없습니다.")
        await session.commit()

    return {"message": "삭제 완료"}


# --- Data API ---
@app.get("/api/data")
async def get_data():
    records = await load_current_records()
    return {"records": records, "total": len(records)}


# --- Upload History ---
@app.get("/api/uploads/history")
async def get_upload_history():
    async with async_session() as session:
        result = await session.execute(
            select(UploadHistory).order_by(UploadHistory.id.desc())
        )
        uploads = [u.to_dict() for u in result.scalars().all()]
    return {"uploads": uploads}


@app.get("/api/uploads/compare")
async def compare_uploads(upload_id_1: int, upload_id_2: int):
    """Compare two uploads: show summary diff, new/removed/changed records."""
    records_1 = await load_records_by_upload(upload_id_1)
    records_2 = await load_records_by_upload(upload_id_2)

    async with async_session() as session:
        r1 = await session.execute(select(UploadHistory).where(UploadHistory.id == upload_id_1))
        r2 = await session.execute(select(UploadHistory).where(UploadHistory.id == upload_id_2))
        h1 = r1.scalar_one_or_none()
        h2 = r2.scalar_one_or_none()
        if not h1 or not h2:
            raise HTTPException(status_code=404, detail="업로드 이력을 찾을 수 없습니다.")
        h1_dict = h1.to_dict()
        h2_dict = h2.to_dict()

    def normalize(text):
        if not text:
            return ""
        return re.sub(r'[.,!?;:\-~·…\s]+', '', text.strip().lower())

    def fingerprint(r):
        return normalize(r.get("content_full", "")) + "|" + normalize(r.get("location", ""))

    def summarize(records):
        total = len(records)
        grades = {"A": 0, "B": 0, "C": 0, "D": 0}
        complete = 0
        for r in records:
            g = r.get("grade_before", "-")
            if g in grades:
                grades[g] += 1
            if r.get("completion") == "완료":
                complete += 1
        rate = round(complete / total * 100, 1) if total > 0 else 0
        return {"total": total, "grades": grades, "complete": complete, "improvement_rate": rate}

    fp1 = {fingerprint(r): r for r in records_1}
    fp2 = {fingerprint(r): r for r in records_2}

    new_records = []
    removed_records = []
    changed_records = []

    for fp, r in fp2.items():
        if fp not in fp1:
            new_records.append({
                "location": r.get("location", ""),
                "content": r.get("content", ""),
                "grade_before": r.get("grade_before", "-"),
                "completion": r.get("completion", ""),
            })
        else:
            old = fp1[fp]
            changes = []
            if old.get("grade_before") != r.get("grade_before"):
                changes.append(f"등급: {old.get('grade_before')} → {r.get('grade_before')}")
            if old.get("completion") != r.get("completion"):
                changes.append(f"완료: {old.get('completion') or '미완료'} → {r.get('completion') or '미완료'}")
            if old.get("grade_after") != r.get("grade_after"):
                changes.append(f"개선후등급: {old.get('grade_after')} → {r.get('grade_after')}")
            if changes:
                changed_records.append({
                    "content": r.get("content", ""),
                    "location": r.get("location", ""),
                    "changes": ", ".join(changes),
                })

    for fp, r in fp1.items():
        if fp not in fp2:
            removed_records.append({
                "location": r.get("location", ""),
                "content": r.get("content", ""),
                "grade_before": r.get("grade_before", "-"),
                "completion": r.get("completion", ""),
            })

    return {
        "upload_1": {**h1_dict, **summarize(records_1)},
        "upload_2": {**h2_dict, **summarize(records_2)},
        "new_records": new_records,
        "removed_records": removed_records,
        "changed_records": changed_records,
    }


# --- Summary ---
@app.get("/api/summary")
async def get_summary(
    request: Request,
    channel: Optional[str] = None,
    year: Optional[str] = None,
    month: Optional[str] = None,
    location: Optional[str] = None,
    grade: Optional[str] = None,
    disaster_type: Optional[str] = None,
    process: Optional[str] = None,
    person: Optional[str] = None,
    week: Optional[int] = None,
    keyword: Optional[str] = None,
    completion: Optional[str] = None,
):
    records = await load_current_records()

    if channel and channel != "전체":
        records = [r for r in records if r.get("channel") == channel]
    if year and year != "전체":
        records = [r for r in records if (r.get("date") or "")[:4] == year]
    if month and month != "전체":
        records = [r for r in records if r["month"] == month]
    if location and location != "전체":
        records = [r for r in records if r["location_group"] == location]
    if grade and grade != "전체":
        records = [r for r in records if r["grade_before"] == grade]
    if disaster_type and disaster_type != "전체":
        records = [r for r in records if r["disaster_type"] == disaster_type]
    if process and process != "전체":
        records = [r for r in records if r["process"] == process]
    if person and person != "전체":
        records = [r for r in records if r["person"] == person]
    if week and week > 0:
        records = [r for r in records if r["week"] == week]
    if completion and completion != "전체":
        records = [r for r in records if r["completion"] == completion]
    if keyword:
        kw = keyword.lower()
        records = [r for r in records if kw in r.get("content_full", "").lower()
                   or kw in r.get("location", "").lower()
                   or kw in r.get("improvement_plan", "").lower()]

    from collections import Counter

    def normalize_content(text):
        if not text:
            return ""
        t = text.strip()
        t = re.sub(r'[.,!?;:\-~·…\s]+', '', t)
        return t.lower()

    content_counts = Counter()
    norm_map = {}
    for r in records:
        c = r.get("content_full", "").strip()
        norm = normalize_content(c)
        if norm:
            content_counts[norm] += 1
            norm_map[id(r)] = norm

    for r in records:
        norm = norm_map.get(id(r), "")
        cnt = content_counts.get(norm, 0)
        r["repeat_count"] = cnt
        r["is_repeat"] = cnt >= 2

    total = len(records)
    repeat_total = sum(1 for r in records if r["is_repeat"])

    grade_a = sum(1 for r in records if r["grade_before"] == "A")
    grade_b = sum(1 for r in records if r["grade_before"] == "B")
    grade_c = sum(1 for r in records if r["grade_before"] == "C")
    grade_d = sum(1 for r in records if r["grade_before"] == "D")

    complete = sum(1 for r in records if r["completion"] == "완료")
    incomplete = sum(1 for r in records if r["completion"] != "완료")
    improvement_rate = round(complete / total * 100, 1) if total > 0 else 0

    def month_sort_key(m):
        try:
            return int(m.replace("월", ""))
        except (ValueError, AttributeError):
            return 0

    all_months = sorted(set(r["month"] for r in records), key=month_sort_key)
    grade_cumulative = {}
    cumul = {"A": 0, "B": 0, "C": 0, "D": 0}
    cumul_total = 0
    cumul_complete = 0
    for m in all_months:
        month_recs = [r for r in records if r["month"] == m]
        for r in month_recs:
            g = r["grade_before"] if r["grade_before"] in ("A", "B", "C", "D") else None
            if g and r["completion"] != "완료":
                cumul[g] += 1
        cumul_total += len(month_recs)
        cumul_complete += sum(1 for r in month_recs if r["completion"] == "완료")
        grade_cumulative[m] = {
            "A": cumul["A"], "B": cumul["B"], "C": cumul["C"], "D": cumul["D"],
            "total_remaining": cumul_total - cumul_complete,
        }

    location_stats: dict[str, dict[str, int]] = {}
    for r in records:
        lg = r["location_group"]
        if lg not in location_stats:
            location_stats[lg] = {"A": 0, "B": 0, "C": 0, "D": 0, "-": 0}
        g = r["grade_before"] if r["grade_before"] in ("A", "B", "C", "D") else "-"
        location_stats[lg][g] += 1

    location_disaster_stats: dict[str, dict[str, int]] = {}
    for r in records:
        lg = r["location_group"]
        dt = r["disaster_type"] if r["disaster_type"] else "미분류"
        if lg not in location_disaster_stats:
            location_disaster_stats[lg] = {}
        location_disaster_stats[lg][dt] = location_disaster_stats[lg].get(dt, 0) + 1

    grade_trend: dict[str, dict[str, int]] = {}
    for r in records:
        m = r["month"]
        if m not in grade_trend:
            grade_trend[m] = {"A": 0, "B": 0, "C": 0, "D": 0}
        g = r["grade_before"] if r["grade_before"] in ("A", "B", "C", "D") else None
        if g:
            grade_trend[m][g] += 1

    def month_sort_key2(m):
        try:
            return int(m.replace("월", ""))
        except (ValueError, AttributeError):
            return m
    grade_trend = dict(sorted(grade_trend.items(), key=lambda x: month_sort_key2(x[0])))

    week_stats: dict[str, int] = {}
    for r in records:
        m = r["month"]
        w = r["week"]
        if w > 0:
            key = f"{m} {w}주차"
            week_stats[key] = week_stats.get(key, 0) + 1

    disaster_stats: dict[str, int] = {}
    for r in records:
        dt = r["disaster_type"] if r["disaster_type"] else "미분류"
        disaster_stats[dt] = disaster_stats.get(dt, 0) + 1

    process_stats: dict[str, int] = {}
    for r in records:
        p = r["process"] if r["process"] else "미분류"
        process_stats[p] = process_stats.get(p, 0) + 1

    channel_stats: dict[str, int] = {}
    for r in records:
        ch = r.get("channel", "미분류")
        channel_stats[ch] = channel_stats.get(ch, 0) + 1

    channel_grade_stats: dict[str, dict[str, int]] = {}
    for r in records:
        ch = r.get("channel", "미분류")
        if ch not in channel_grade_stats:
            channel_grade_stats[ch] = {"A": 0, "B": 0, "C": 0, "D": 0, "-": 0, "complete": 0, "incomplete": 0}
        g = r["grade_before"] if r["grade_before"] in ("A", "B", "C", "D") else "-"
        channel_grade_stats[ch][g] += 1
        if r["completion"] == "완료":
            channel_grade_stats[ch]["complete"] += 1
        else:
            channel_grade_stats[ch]["incomplete"] += 1

    all_records = await load_current_records()
    channels = sorted(set(r.get("channel", "미분류") for r in all_records))
    years = sorted(set(r["date"][:4] for r in all_records if r.get("date") and len(r["date"]) >= 4))
    months = sorted(set(r["month"] for r in all_records))
    locations = sorted(set(r["location_group"] for r in all_records))
    disaster_types = sorted(set(r["disaster_type"] for r in all_records if r["disaster_type"]))
    processes = sorted(set(r["process"] for r in all_records if r["process"]))
    persons = sorted(set(r["person"] for r in all_records if r["person"]))
    weeks = sorted(set(r["week"] for r in all_records if r["week"] > 0))

    return {
        "total": total,
        "improvement_rate": improvement_rate,
        "repeat_total": repeat_total,
        "grade_a": grade_a,
        "grade_b": grade_b,
        "grade_c": grade_c,
        "grade_d": grade_d,
        "grade_cumulative": grade_cumulative,
        "complete": complete,
        "incomplete": incomplete,
        "location_stats": location_stats,
        "location_disaster_stats": location_disaster_stats,
        "grade_trend": grade_trend,
        "week_stats": week_stats,
        "disaster_stats": disaster_stats,
        "process_stats": process_stats,
        "channel_stats": channel_stats,
        "channel_grade_stats": channel_grade_stats,
        "records": records,
        "filters": {
            "channels": channels,
            "years": years,
            "months": months,
            "locations": locations,
            "disaster_types": disaster_types,
            "processes": processes,
            "persons": persons,
            "weeks": weeks,
        },
    }


# --- Static Files ---
app.mount("/static", StaticFiles(directory="static"), name="static")


@app.get("/")
async def root():
    return FileResponse("static/index.html")


if __name__ == "__main__":
    import uvicorn
    port = int(os.environ.get("PORT", 8000))
    uvicorn.run(app, host="0.0.0.0", port=port)
